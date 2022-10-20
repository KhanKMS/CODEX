# 파일명 지정
template_file = './input/invoice-template.xlsx'
save_file = './output/invoice_pc.xlsx'

# 오늘 날짜 구하기 
from datetime import datetime
date = datetime.today()

#입력 데이터 
name = '강면수'
subject = '10월 청구분'
items = [ # 청구 내역
    ['CPU', 11, 320000],
    ['메모리', 7, 88000], 
    ['SSD', 4, 66000],
    ['HDD', 7, 45000],
    ['GPU', 3, 550000],
    ['파워', 12, 67000], 
    ['케이스', 15, 55000],
    ['키보드', 9, 15000],
    ['마우스', 8, 9000]
]

# 템플릿 열기 
import openpyxl as excel
book = excel.load_workbook(template_file)
sheet = book.active 

# 청구서 발행일 입력 
sheet["G3"] = date.strftime("%Y-%m-%d") # 날짜를 2022-10-01형태로 파싱

# 템플릿에 성명과 청구 건명 입력 
sheet["B4"] = name
sheet["C10"] = subject

# 거래 내역 채우기 
total = 0 
for i, it in enumerate(items): 
    summary, count, price = it
    subtotal = count * price
    total += subtotal
    # 시트에 채우기 
    row = 15+i
    sheet.cell(row, 2, summary) 
    sheet.cell(row, 5, count)     
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)

# 청구 금액 입력 
sheet['C11'] = total

# 파일 저장
book.save(save_file)