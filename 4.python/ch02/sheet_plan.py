import openpyxl as excel 

book = excel.load_workbook("input/all-customer.xlsx")
sheet = book["명부"]

for row in sheet.iter_rows(min_row=3):
    cells = [v.value for v in row]
    if cells[0] is None:break
    print(cells)
    (name, area, plan) = cells 
    sname = plan+"플랜"
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["이름", "주소", "플랜"])
    else:
        to_sheet = book[sname]
    to_sheet.append(cells)

book.save("output/sheet_plan.xlsx")