import openpyxl as excel 

book = excel.load_workbook("input/auditlist.xlsx")
sheet = book["career"]

for row in sheet.iter_rows(min_row=2):
    cells = [v.value for v in row]
    if cells[0] is None:break
    print(cells)
    (proj,job,org) = cells 
    oname = org+"발주기관"
    if oname not in book.sheetnames:
        to_sheet = book.create_sheet(title=oname)
        to_sheet.append(["참여사업명","업무","발주처"])
    else:
        to_sheet = book[oname]
    to_sheet.append(cells)
    
book.save("output/audit_list_sep.xlsx")