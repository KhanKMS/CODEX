import openpyxl as excel 
import datetime

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "출생기간"
sheet["B1"] = "초등학교 입학 연도"
sheet["C1"] = "대학교 학번"

sheet.column_dimensions['A'].width=40
sheet.column_dimensions['B'].width=20
sheet.column_dimensions['C'].width=20

for i in range(50): # range는 0부터 49까지 50번 반복 후 빠져나간다.
    birth_year = 2002-i 
    birth_range = "{}년 3월 1일생 ~ {}년 2월 28(29)일생".format(birth_year, birth_year+1)

    ele_year = birth_year + 7 # 태어난 후 7년 뒤 초등학교 입학 
    univ_year = birth_year + 19 # 태어난 후 19년 후 대학교 입학 
    univ_num = str(univ_year)[2:] 
    # str() 메소드 : 입학년도 4자리를 문자열로 변환 
    # [2:] index 2 부터 마지막 문자까지 가져온다. 
    # [2012] 에서 index 2부터 끝까지 12 가져옴. (0, 1, 2, 3)  

    sheet.cell(i+2, 1, value=birth_range)
    sheet.cell(i+2, 2, value=str(ele_year)+"년")
    sheet.cell(i+2, 3, value=univ_num+"학번")

sheet["A2"] = "2002년 3월 1일생 ~ 2002년 12월 31일생"

book.save("output/write_entry_year.xlsx")