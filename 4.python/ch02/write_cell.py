# openpyxl 모듈을 가져오고, 이후 excel 모듈로 명명함 
import openpyxl as excel 

# 워크북을 생성 
book = excel.Workbook()

# 활성화된 워크시트를 sheet 변수에 저장 
sheet = book.active 

# A1에 값을 설정함 
sheet["A1"] = "일찍 일어나는 새가 벌레를 잡는다"

# 두번째 열(row=2), 첫번째 행(column=1)에 값을 설정함. 즉 A2  
sheet.cell(row=2, column=1, value="하늘은 스스로 돕는 자를 돕는다")

# 세번째 열(row=3), 첫번째 행(column=1)에 값을 설정함. 즉 A3
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "낙숫물이 바위를 뚫는다"

# 워크북 저장 
book.save("output/write_cell.xlsx")