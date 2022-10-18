import openpyxl as excel 

book = excel.load_workbook("input/auditlist.xlsx")
sheet = book["career"]

career = [["번호", "년도","수행기간","참여사업명","업무","발주처","출장","KOSA","파트","회사"]]

for row in sheet.iter_rows(min_row=2):
    values = [v.value for v in row]
    if values[0] is None: break 
    site = values[5]
    trip = values[6]
    if site == "국민건강보험공단" and trip != None :
        career.append(values)
        print(values)

new_book = excel.Workbook()
new_sheet = new_book.active 
new_sheet["A1"] = "국민건강보험공단 출장 사업 수행"

for row, row_val in enumerate(career):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(2+row, 1+col)
        c.value = val 

new_book.save("output/sheet_career.xlsx")