import openpyxl as excel 
# 엑셀 파일을 읽음, 수식 함수 사용, 활성화된 시트를 가져옴. 
sheet = excel.load_workbook(
    "input/monthly_sales2.xlsx", data_only=True).active
# iter_rows를 사용해 모든 데이터 가져오기 
for row in sheet.iter_rows(min_row=3):
    values=[cell.value for cell in row]
    if values[0] is None : break
    print(values)