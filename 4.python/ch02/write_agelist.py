import openpyxl as excel 
import datetime

book = excel.Workbook()
sheet = book.active

thisyear = datetime.datetime.now().year

sheet["A1"] = "출생연도"
sheet["B1"] = "한국 나이"
sheet["C1"] = "만 나이 (생일 후)"
sheet["D1"] = "만 나이 (생일 전)"

sheet.column_dimensions['C'].width=20
sheet.column_dimensions['D'].width=20

for i in range(80):
    birth_year = thisyear - i 
    korean_age = thisyear - birth_year + 1
    man_age = { 'after_bday':korean_age-1, 'before_bday':korean_age-2}

    year_cell = sheet.cell(i+2, 1)
    year_cell.value = str(birth_year) + "년생"

    age_cell = sheet.cell(i+2, 2)
    age_cell.value = str(korean_age) + "세"

    age_cell = sheet.cell(i+2, 3) 
    age_cell.value = "만" + str(man_age['after_bday']) + "세"

    age_cell = sheet.cell(i+2, 4) 
    age_cell.value = "만" + str(man_age['before_bday']) + "세"

sheet['D2'] = "-"

book.save("output/write2_agelist.xlsx")