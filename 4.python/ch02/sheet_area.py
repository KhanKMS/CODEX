import openpyxl as excel 

book = excel.load_workbook("input/all-customer.xlsx")

sheet = book["명부"]

customers = [["이름", "주소","구매 플랜"]]

for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    if values[0] is None: break 
    area = values[1]
    if area == "서울" or area == "경기도" or area == "인천":
        customers.append(values)
        print(values)

new_book = excel.Workbook()
new_sheet = new_book.active 
new_sheet["A1"] = "수도권 고객 명단"

for row, row_val in enumerate(customers):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(2+row, 1+col)
        c.value = val 

new_book.save("output/sheet_area.xlsx")