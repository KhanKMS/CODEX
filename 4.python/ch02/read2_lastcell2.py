import openpyxl as excel 

sheet = excel.load_workbook("input/monthly_sales2.xlsx").active
print((sheet.max_row, sheet.max_column))