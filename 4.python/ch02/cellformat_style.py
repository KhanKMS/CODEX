import openpyxl as xl 

book = xl.Workbook()
sheet = book.active

sheet.column_dimensions['A'].width = 50
sheet.row_dimensions[1].height = 30

cell = sheet["A1"]
cell.value = "파이썬은 엑셀 자동화 최적 도구입니다."

from openpyxl.styles.alignment import Alignment
cell.alignment = Alignment (
    horizontal='center',    # 수평 정렬 : 가운데
    vertical='center'       # 수직 정렬 : 가운데 
)

from openpyxl.styles.borders import Border, Side
cell.border = Border(
    top=Side(style='thick', color='000000'),     # 실선 종류: thick, thin, medium, double 중 택일 
    right=Side(style='thick', color='000000'),   # 점선 종류: dashed, dotted...
    bottom=Side(style='thick', color='000000'),  # 색상은 00000 RGB 코드 값 --> 흰색 
    left=Side(style='thick', color='000000')     # 000000 RGB 코드 값 --> 검정 
)

from openpyxl.styles import Font
cell.font = Font(
    name='휴먼편지체',
    size=16,        # 폰트 사이즈 16
    bold=True,      # 볼드 사용
    italic=False,   # 이탤릭 해제
    color='FFFFFF'  # FFFFFF RGB 코드 값 --> 흰색
)

from openpyxl.styles import PatternFill
cell.fill = PatternFill(
        fill_type='solid', 
        fgColor='0000FF' # 0000FF RGB 코드 값 --> 파랑
    )

book.save("output/cellformat_style.xlsx")