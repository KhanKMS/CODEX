import openpyxl as excel 
import random

win_num = random.randint(1,100)

book = excel.Workbook()
book.active["B2"] = "'당첨'이라고 적힌 시트를 찾아보자"

for i in range(1, 101):
    sname = str(i) + "번"
    sheet = book.create_sheet(title=sname)
    word = "꽝"
    if i  == win_num: word="당첨"
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1, x+1)
            c.value = word

book.save("./output/sheet_game.xlsx")
print("ok, winning number =", win_num )